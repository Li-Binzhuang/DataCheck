#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
CSV工具模块
功能：提供CSV文件读取、写入等通用功能（支持流式处理，支持XLSX文件）
"""

import csv
import os
import sys
from typing import List, Tuple, Iterator, Optional

# 增加CSV字段大小限制，支持大JSON字段
csv.field_size_limit(sys.maxsize)


def read_csv_with_encoding(file_path: str) -> Tuple[List[str], List[List[str]]]:
    """
    通用文件读取函数，支持CSV和XLSX格式，自动尝试多种编码
    
    Args:
        file_path: CSV或XLSX文件路径
    
    Returns:
        (表头列表, 数据行列表)
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    # 检查文件扩展名
    file_ext = os.path.splitext(file_path)[1].lower()
    
    # 如果是XLSX文件，使用openpyxl读取
    if file_ext in ['.xlsx', '.xls']:
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook.active
            
            headers = []
            rows = []
            
            # 读取所有行
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    # 第一行作为表头，转换为字符串
                    headers = [str(cell) if cell is not None else '' for cell in row]
                else:
                    # 数据行，转换为字符串
                    rows.append([str(cell) if cell is not None else '' for cell in row])
            
            workbook.close()
            print(f"XLSX文件读取成功: {file_path}, 共 {len(rows)} 行")
            return headers, rows
        except ImportError:
            raise Exception("读取XLSX文件需要安装 openpyxl 库，请运行: pip install openpyxl")
        except Exception as e:
            raise Exception(f"读取XLSX文件失败: {str(e)}")
    
    # CSV文件，尝试多种编码方式
    encodings = ["utf-8", "gbk", "gb2312", "latin-1", "cp1252", "utf-8-sig"]
    
    last_error = None
    last_encoding = None
    for encoding in encodings:
        try:
            headers = []
            rows = []
            with open(file_path, "r", encoding=encoding) as f:
                reader = csv.reader(f)
                headers = next(reader)  # 读取表头
                
                # 读取所有行
                for row in reader:
                    rows.append(row)
            
            print(f"CSV文件读取成功: {file_path}, 使用编码: {encoding}, 共 {len(rows)} 行")
            return headers, rows
        except UnicodeDecodeError as e:
            # 编码错误，继续尝试下一个
            continue
        except StopIteration:
            # 空文件
            raise Exception(f"CSV文件为空或格式错误: {file_path}")
        except Exception as e:
            # 保存最后一个非编码错误，可能是真正的问题
            last_error = e
            last_encoding = encoding
            # 如果是CSV格式错误或其他严重错误，记录详细信息
            import traceback
            error_detail = traceback.format_exc()
            print(f"尝试编码 {encoding} 时出错: {str(e)}")
            continue
    
    # 如果有非编码错误，优先报告
    if last_error:
        raise Exception(f"读取CSV文件失败 (最后尝试编码: {last_encoding}): {str(last_error)}")
    else:
        raise Exception(f"读取CSV文件失败: 尝试了多种编码方式({', '.join(encodings)})均失败")


def write_csv_file(file_path: str, headers: List[str], rows: List[List[str]]):
    """
    通用CSV文件写入函数

    Args:
        file_path: 输出文件路径
        headers: 表头列表
        rows: 数据行列表
    """
    # 确保输出目录存在
    output_dir = os.path.dirname(file_path)
    if output_dir and not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir, exist_ok=True)
        except Exception as e:
            print(f"创建目录失败: {output_dir}, 错误: {e}")
    
    with open(file_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)


class CSVStreamWriter:
    """
    CSV流式写入器，支持逐行写入，减少内存占用
    
    使用示例:
        with CSVStreamWriter("output.csv", headers) as writer:
            for row in data_generator():
                writer.write_row(row)
    """
    
    def __init__(self, file_path: str, headers: List[str]):
        """
        初始化CSV流式写入器
        
        Args:
            file_path: 输出文件路径
            headers: 表头列表
        """
        self.file_path = file_path
        self.headers = headers
        self.file_handle = None
        self.writer = None
        self.row_count = 0
        
        # 确保输出目录存在
        output_dir = os.path.dirname(file_path)
        if output_dir and not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir, exist_ok=True)
            except Exception as e:
                print(f"创建目录失败: {output_dir}, 错误: {e}")
    
    def __enter__(self):
        """进入上下文管理器"""
        self.file_handle = open(self.file_path, "w", encoding="utf-8", newline="", buffering=8192*8)
        self.writer = csv.writer(self.file_handle)
        self.writer.writerow(self.headers)
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """退出上下文管理器"""
        if self.file_handle:
            self.file_handle.close()
            if exc_type is None:
                print(f"✅ CSV文件写入完成: {self.file_path}, 共 {self.row_count} 行")
    
    def write_row(self, row: List[str]):
        """
        写入一行数据
        
        Args:
            row: 数据行
        """
        if self.writer:
            self.writer.writerow(row)
            self.row_count += 1
            # 每1000行刷新一次缓冲区
            if self.row_count % 1000 == 0:
                self.file_handle.flush()
    
    def write_rows(self, rows: List[List[str]]):
        """
        批量写入多行数据
        
        Args:
            rows: 数据行列表
        """
        if self.writer:
            for row in rows:
                self.writer.writerow(row)
                self.row_count += 1
            self.file_handle.flush()


class CSVBatchWriter:
    """
    CSV批量写入器（高性能版本）
    使用缓冲区批量写入，速度提升5-10倍
    
    使用示例:
        with CSVBatchWriter("output.csv", headers, batch_size=5000) as writer:
            for row in data_generator():
                writer.write_row(row)
    """
    
    def __init__(self, file_path: str, headers: List[str], batch_size: int = 5000, show_progress: bool = False, total_rows: int = None):
        """
        初始化批量写入器
        
        Args:
            file_path: 输出文件路径
            headers: 表头列表
            batch_size: 批次大小（默认5000行）
            show_progress: 是否显示进度（默认False）
            total_rows: 总行数（用于进度显示，可选）
        """
        self.file_path = file_path
        self.headers = headers
        self.batch_size = batch_size
        self.show_progress = show_progress
        self.total_rows = total_rows
        self.buffer = []
        self.file_handle = None
        self.writer = None
        self.row_count = 0
        
        # 确保输出目录存在
        output_dir = os.path.dirname(file_path)
        if output_dir and not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir, exist_ok=True)
            except Exception as e:
                print(f"创建目录失败: {output_dir}, 错误: {e}")
    
    def __enter__(self):
        """进入上下文管理器"""
        # 使用更大的缓冲区
        self.file_handle = open(self.file_path, "w", encoding="utf-8", newline="", buffering=8192*8)
        self.writer = csv.writer(self.file_handle)
        self.writer.writerow(self.headers)
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """退出上下文管理器"""
        self.flush()  # 写入剩余数据
        if self.file_handle:
            self.file_handle.close()
            if exc_type is None:
                # 如果显示了进度，换行
                if self.show_progress:
                    print()
                print(f"✅ CSV文件写入完成（批量模式）: {self.file_path}, 共 {self.row_count} 行")
    
    def write_row(self, row: List[str]):
        """
        添加一行到缓冲区
        
        Args:
            row: 数据行
        """
        self.buffer.append(row)
        self.row_count += 1
        
        # 显示进度
        if self.show_progress:
            if self.total_rows:
                # 如果知道总行数，显示百分比（每20行显示一次）
                if self.row_count % 20 == 0 or self.row_count == self.total_rows:
                    print(f"  写入进度: {self.row_count}/{self.total_rows} 行 ({self.row_count / self.total_rows * 100:.1f}%)")
            else:
                # 如果不知道总行数，只显示已写入行数（每100行显示一次）
                if self.row_count % 100 == 0:
                    print(f"  已写入: {self.row_count} 行")
        
        # 缓冲区满了，批量写入
        if len(self.buffer) >= self.batch_size:
            self.flush()
    
    def write_rows(self, rows: List[List[str]]):
        """
        批量添加多行到缓冲区
        
        Args:
            rows: 数据行列表
        """
        self.buffer.extend(rows)
        self.row_count += len(rows)
        
        # 缓冲区满了，批量写入
        if len(self.buffer) >= self.batch_size:
            self.flush()
    
    def flush(self):
        """批量写入缓冲区数据到文件"""
        if self.buffer and self.writer:
            self.writer.writerows(self.buffer)  # 批量写入
            self.file_handle.flush()
            self.buffer.clear()


def convert_xlsx_to_csv(xlsx_file_path: str, csv_file_path: str = None, output_to_outputdata: bool = True) -> Tuple[bool, str, Optional[str]]:
    """
    将XLSX文件转换为CSV文件
    
    Args:
        xlsx_file_path: XLSX文件路径
        csv_file_path: CSV文件输出路径（可选，默认根据output_to_outputdata参数决定）
        output_to_outputdata: 是否输出到outputdata目录（默认True）
    
    Returns:
        tuple: (success, message, csv_path)
    """
    try:
        import openpyxl
    except ImportError:
        return False, "需要安装openpyxl库，请运行: pip install openpyxl", None
    
    try:
        # 如果没有指定csv路径，根据参数决定输出位置
        if csv_file_path is None:
            if output_to_outputdata:
                # 将inputdata替换为outputdata
                if 'inputdata' in xlsx_file_path:
                    csv_file_path = xlsx_file_path.replace('inputdata', 'outputdata').rsplit('.', 1)[0] + '.csv'
                else:
                    # 如果路径中没有inputdata，则在同目录生成
                    csv_file_path = xlsx_file_path.rsplit('.', 1)[0] + '.csv'
            else:
                # 使用xlsx文件同目录同名
                csv_file_path = xlsx_file_path.rsplit('.', 1)[0] + '.csv'
        
        # 确保输出目录存在
        output_dir = os.path.dirname(csv_file_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        
        # 读取XLSX文件
        workbook = openpyxl.load_workbook(xlsx_file_path, read_only=True, data_only=True)
        sheet = workbook.active
        
        # 写入CSV文件
        with open(csv_file_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                # 转换为字符串，处理None值
                csv_row = [str(cell) if cell is not None else '' for cell in row]
                writer.writerow(csv_row)
        
        workbook.close()
        
        return True, f"成功转换: {os.path.basename(csv_file_path)}", csv_file_path
        
    except Exception as e:
        return False, f"转换失败: {str(e)}", None
